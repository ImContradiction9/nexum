"""Exportação de transações e extratos em Excel (.xlsx) via openpyxl."""
import os
import subprocess
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response, JSONResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..deps import get_db
from ..rede import eh_local
from .transacoes import listar_transacoes
from .extrato import extrato_conta

router = APIRouter()

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FMT_MOEDA = 'R$ #,##0.00'
_FMT_DATA = "DD/MM/YYYY"


def _data(iso):
    """ISO 'YYYY-MM-DD' → date (pro Excel formatar como data)."""
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso).date()
    except (ValueError, TypeError):
        return iso


def _planilha(titulo, colunas, linhas):
    """Monta um Workbook com cabeçalho em negrito, congelado, e auto-largura.
    `colunas` = [(titulo, fmt|None)]; `linhas` = lista de tuplas (mesmo tamanho)."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]
    cab_fill = PatternFill("solid", fgColor="1F2937")
    cab_font = Font(bold=True, color="FFFFFF")
    for c, (nome, _fmt) in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=c, value=nome)
        cel.font = cab_font
        cel.fill = cab_fill
        cel.alignment = Alignment(horizontal="center")
    for r, linha in enumerate(linhas, start=2):
        for c, valor in enumerate(linha, start=1):
            cel = ws.cell(row=r, column=c, value=valor)
            fmt = colunas[c - 1][1]
            if fmt and valor is not None:
                cel.number_format = fmt
    ws.freeze_panes = "A2"
    # Auto-largura simples
    for c, (nome, _fmt) in enumerate(colunas, start=1):
        largura = len(str(nome))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                largura = max(largura, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(largura + 2, 10), 48)
    return wb


def _pasta_downloads() -> Path:
    """Pasta onde salvar os exports: <Downloads>/Nexum (fallback: home/Nexum)."""
    home = Path.home()
    base = home / "Downloads"
    if not base.exists():
        base = home
    pasta = base / "Nexum"
    try:
        pasta.mkdir(parents=True, exist_ok=True)
    except Exception:
        pasta = base
    return pasta


def _abrir_no_explorer(caminho: Path):
    """Abre o Explorer com o arquivo selecionado (Windows). Best-effort.
    Não roda em modo headless (NEXUM_NO_BROWSER=1) pra não abrir janela em testes."""
    if os.environ.get("NEXUM_NO_BROWSER") == "1":
        return
    try:
        if sys.platform.startswith("win"):
            subprocess.Popen(["explorer", "/select,", str(caminho)])
        elif sys.platform == "darwin":
            subprocess.Popen(["open", "-R", str(caminho)])
        else:
            subprocess.Popen(["xdg-open", str(caminho.parent)])
    except Exception:
        pass


def _entregar(wb, nome_base: str, request: Request):
    """Acesso LOCAL (janela do PC): salva o .xlsx na pasta Downloads/Nexum e abre
    o Explorer (a janela nativa não tem 'baixar arquivo' visível). Acesso REMOTO
    (celular/navegador): devolve o arquivo como download."""
    host = request.client.host if (request and request.client) else None
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    nome_arquivo = f"{nome_base}_{ts}.xlsx"
    if eh_local(host):
        pasta = _pasta_downloads()
        caminho = pasta / nome_arquivo
        wb.save(str(caminho))
        _abrir_no_explorer(caminho)
        return JSONResponse({"ok": True, "arquivo": str(caminho), "pasta": str(pasta)})
    # Remoto → download via HTTP
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/api/exportar/transacoes")
def exportar_transacoes(
    request: Request,
    mes: Optional[str] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    conta_id: Optional[int] = None,
    banco_id: Optional[int] = None,
    tipo_conta: Optional[str] = None,
    categoria_id: Optional[int] = None,
    atribuicao_id: Optional[int] = None,
    busca: Optional[str] = None,
    nao_categorizado: bool = False,
    nao_atribuido: bool = False,
    incluir_transferencias: bool = False,
    incluir_suspeitas: bool = False,
    db: Session = Depends(get_db),
):
    """Exporta as transações (respeitando os mesmos filtros da listagem) em .xlsx."""
    res = listar_transacoes(
        mes=mes, data_inicio=data_inicio, data_fim=data_fim, conta_id=conta_id,
        banco_id=banco_id, tipo_conta=tipo_conta, categoria_id=categoria_id,
        atribuicao_id=atribuicao_id, busca=busca, nao_categorizado=nao_categorizado,
        nao_atribuido=nao_atribuido, incluir_transferencias=incluir_transferencias,
        incluir_suspeitas=incluir_suspeitas, skip=0, limit=1000000, db=db,
    )
    colunas = [
        ("Data", _FMT_DATA), ("Descrição", None), ("Descrição personalizada", None),
        ("Valor (R$)", _FMT_MOEDA), ("Tipo", None), ("Conta", None), ("Banco", None),
        ("Categoria", None), ("Atribuição", None), ("Forma de pagamento", None),
        ("Parcela", None), ("Mês ref.", None), ("Observações", None),
    ]
    linhas = []
    for t in res["items"]:
        # Valor com sinal: Receita + ; Despesa e abatedora - (facilita somar no Excel)
        v = t["valor"] or 0
        valor_signed = v if (t["tipo"] == "Receita" and not t.get("eh_abatedora")) else -v
        linhas.append((
            _data(t["data"]), t["descricao"], t.get("descricao_personalizada") or "",
            round(valor_signed, 2), t["tipo"], t.get("conta") or "", t.get("banco") or "",
            t.get("categoria") or "", t.get("atribuicao") or "", t.get("forma_pagamento") or "",
            t.get("parcela") or "", t.get("mes_referencia") or "", t.get("observacoes") or "",
        ))
    wb = _planilha("Transações", colunas, linhas)
    return _entregar(wb, "transacoes", request)


@router.get("/api/exportar/extrato")
def exportar_extrato(
    request: Request,
    conta_id: int,
    mes: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Exporta o extrato de uma conta (saldo corrido) em .xlsx."""
    ext = extrato_conta(conta_id=conta_id, mes=mes, db=db)
    conta = ext.get("conta") or {}
    colunas = [
        ("Data", _FMT_DATA), ("Descrição", None), ("Categoria", None), ("Atribuição", None),
        ("Tipo", None), ("Entrada/Saída (R$)", _FMT_MOEDA), ("Saldo após (R$)", _FMT_MOEDA),
    ]
    linhas = []
    # Linha do saldo inicial
    primeira = ext["items"][0]["data"] if ext.get("items") else None
    linhas.append((_data(primeira), "Saldo inicial", "", "", "",
                   None, round(ext.get("saldo_inicial") or 0, 2)))
    for i in ext.get("items", []):
        linhas.append((
            _data(i["data"]), i["descricao"], i.get("categoria") or "", i.get("atribuicao") or "",
            i["tipo"], round(i.get("delta") or 0, 2), round(i.get("saldo_apos") or 0, 2),
        ))
    titulo = (conta.get("nome") or "Extrato")
    wb = _planilha(titulo, colunas, linhas)
    nome_base = "extrato_" + (conta.get("nome") or "conta").replace(" ", "_")
    return _entregar(wb, nome_base, request)
