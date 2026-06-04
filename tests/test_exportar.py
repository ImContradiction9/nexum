"""Exportação .xlsx de transações e extrato."""
from datetime import date
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.routers.exportar import exportar_transacoes, exportar_extrato, XLSX_MEDIA
from app.database import Transacao, Categoria, Conta


# Request "remoto" (host não-local) → endpoint devolve o arquivo p/ download,
# sem salvar em disco (o caminho local salvaria em Downloads, efeito colateral).
def _req_remoto():
    return SimpleNamespace(client=SimpleNamespace(host="192.168.0.99"))


def _conta(db, tipo="Conta Corrente"):
    c = Conta(nome="Nubank Conta", tipo=tipo)
    db.add(c); db.commit(); db.refresh(c)
    return c


def _t(db, conta, descricao, valor, tipo="Despesa", cat=None):
    db.add(Transacao(
        conta_id=conta.id, data=date(2026, 5, 10), descricao=descricao,
        descricao_normalizada=descricao.lower(), valor=valor, tipo=tipo,
        mes_referencia="05/2026", categoria_id=cat, categoria_origem="manual",
    ))
    db.commit()


def _wb(resp):
    assert resp.media_type == XLSX_MEDIA
    assert "attachment" in resp.headers["content-disposition"]
    return load_workbook(BytesIO(resp.body))


def test_exportar_transacoes_gera_xlsx(db):
    conta = _conta(db)
    cat = Categoria(nome="Mercado", tipo="Despesa")
    db.add(cat); db.commit(); db.refresh(cat)
    _t(db, conta, "MERCADO XPTO", 50.0, "Despesa", cat.id)
    _t(db, conta, "SALARIO", 1000.0, "Receita")

    resp = exportar_transacoes(request=_req_remoto(), db=db)
    wb = _wb(resp)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Data"      # cabeçalho
    # 2 transações = 3 linhas (cabeçalho + 2)
    assert ws.max_row == 3
    # Valor com sinal: despesa negativa, receita positiva
    valores = [ws.cell(row=r, column=4).value for r in (2, 3)]
    assert -50.0 in valores and 1000.0 in valores


def test_exportar_extrato_gera_xlsx_com_saldo(db):
    conta = _conta(db)
    _t(db, conta, "PIX RECEBIDO", 200.0, "Receita")
    _t(db, conta, "COMPRA", 80.0, "Despesa")

    resp = exportar_extrato(request=_req_remoto(), conta_id=conta.id, mes="05/2026", db=db)
    wb = _wb(resp)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Data"
    # 1ª linha de dados = "Saldo inicial"
    assert ws.cell(row=2, column=2).value == "Saldo inicial"
    # cabeçalho + saldo inicial + 2 transações = 4 linhas
    assert ws.max_row == 4


def test_exportar_local_salva_em_disco(db, tmp_path, monkeypatch):
    """Acesso local (PC): salva o .xlsx na pasta e devolve JSON com o caminho."""
    import json
    from pathlib import Path
    import app.routers.exportar as ex
    monkeypatch.setattr(ex, "_pasta_downloads", lambda: tmp_path)
    monkeypatch.setenv("NEXUM_NO_BROWSER", "1")   # não abre o Explorer no teste
    conta = _conta(db)
    _t(db, conta, "COMPRA", 10.0)
    req = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
    resp = exportar_transacoes(request=req, db=db)
    data = json.loads(bytes(resp.body))
    assert data["ok"] is True
    arq = Path(data["arquivo"])
    assert arq.exists() and arq.suffix == ".xlsx"
    assert arq.parent == tmp_path
